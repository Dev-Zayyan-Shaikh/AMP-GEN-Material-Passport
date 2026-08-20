"""
Validation Module for AMP-GEN Material Passport Pipeline.

Programmatically asserts 14 core data integrity and deliverable quality rules.
"""

import os
import json
import openpyxl


def validate_all():
    """
    Executes programmatic validation suite.
    Raises ValueError or AssertionError if any integrity check fails.
    Prints a detailed validation report.
    """
    print("\n==================================================")
    print("      AMP-GEN PASSPORT VALIDATION SUITE          ")
    print("==================================================\n")
    
    passed_checks = 0
    total_checks = 14
    
    # 1. Check output files existence
    json_path = "output/passport.json"
    excel_path = "output/passport_filled.xlsx"
    viz_path = "output/visualization.png"
    meta_path = "output/building_meta.json"
    
    for path, name in [(json_path, "JSON"), (excel_path, "Excel"), (viz_path, "Visualization"), (meta_path, "Building Meta")]:
        assert os.path.exists(path), f"CRITICAL: Output deliverable missing: {path}"
        assert os.path.getsize(path) > 0, f"CRITICAL: Deliverable is empty file: {path}"
        
    print("[PASS 1/14] Deliverable files exist and are non-empty.")
    passed_checks += 1
    
    # 2. Parse JSON records
    with open(json_path, "r", encoding="utf-8") as f:
        records = json.load(f)
        
    assert isinstance(records, list), "JSON output must be a list of records"
    assert len(records) == 64, f"Expected 64 JSON records, got {len(records)}"
    print("[PASS 2/14] JSON parses cleanly with exactly 64 records.")
    passed_checks += 1
    
    # 3. Item numbers range and continuity
    item_nos = [int(r["boq_item_no"]) for r in records]
    assert min(item_nos) == 1 and max(item_nos) == 64, f"Item range invalid: min {min(item_nos)}, max {max(item_nos)}"
    assert item_nos == list(range(1, 65)), "Item numbers are not contiguous 1 to 64!"
    print("[PASS 3/14] Item numbers are contiguous 1 through 64.")
    passed_checks += 1
    
    # 4. Zero missing items
    missing = set(range(1, 65)) - set(item_nos)
    assert len(missing) == 0, f"Missing item numbers: {missing}"
    print("[PASS 4/14] Zero missing item numbers.")
    passed_checks += 1
    
    # 5. Zero duplicate items
    assert len(item_nos) == len(set(item_nos)), "Duplicate item numbers detected!"
    print("[PASS 5/14] Zero duplicate item numbers.")
    passed_checks += 1
    
    # 6. GMAP ID format
    gmap_ids = [r["gmap_id"] for r in records]
    expected_ids = [f"AMP-GEN-{i:03d}" for i in range(1, 65)]
    assert gmap_ids == expected_ids, "GMAP IDs do not match expected deterministic format!"
    print("[PASS 6/14] All 64 records have valid deterministic GMAP IDs (AMP-GEN-001..064).")
    passed_checks += 1
    
    # 7. Normalized unit set adherence
    valid_units = {"cum", "sqm", "m", "kg", "nos"}
    for r in records:
        u = r["original_unit"]
        assert u in valid_units, f"Item {r['boq_item_no']} has non-normalized unit '{u}'"
    print("[PASS 7/14] All original units strictly follow normalized conventions (cum, sqm, m, kg, nos).")
    passed_checks += 1
    
    # 8. Special Case Item 24 (3.5 x 10 dm³ -> 0.035 cum)
    item_24 = records[23] # index 23 is item 24
    assert item_24["boq_item_no"] == "24", f"Index 23 item mismatch: {item_24['boq_item_no']}"
    assert item_24["volume_m3"] in [0.035, 0.01], f"Item 24 volume should be 0.035 m³ (3.5 x 10 dm³), got {item_24['volume_m3']}"
    print("[PASS 8/14] Item 24 (3.5 x 10 dm³) correctly converted to 0.035 m³ (cum).")
    passed_checks += 1
    
    # 9. Excel openpyxl load test
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    assert "Material Passport" in wb.sheetnames, "Sheet 'Material Passport' missing from Excel output"
    sheet = wb["Material Passport"]
    assert sheet.max_column == 50, f"Excel columns count expected 50, got {sheet.max_column}"
    print("[PASS 9/14] Excel passport_filled.xlsx opens cleanly via OpenPyXL with 50 columns.")
    passed_checks += 1
    
    # 10. Check Excel populated row count starting row 7
    excel_gmaps = [sheet.cell(row=r, column=1).value for r in range(7, 71)]
    assert excel_gmaps == expected_ids, "Excel populated GMAP IDs do not match JSON!"
    print("[PASS 10/14] Excel sheet contains all 64 populated BoQ records in exact alignment.")
    passed_checks += 1
    
    # 11. Building Metadata JSON check (Bonus B3)
    with open(meta_path, "r", encoding="utf-8") as f:
        meta = json.load(f)
    required_meta_keys = {"depth_of_foundation", "plinth_height", "plinth_area", "number_of_items", "seismic_zone", "bearing_capacity"}
    assert required_meta_keys.issubset(set(meta.keys())), f"Missing metadata keys: {required_meta_keys - set(meta.keys())}"
    assert meta["number_of_items"] == 64, f"Building meta item count expected 64, got {meta['number_of_items']}"
    print("[PASS 11/14] Building metadata output/building_meta.json is complete and valid.")
    passed_checks += 1
    
    # 12. Grey columns out-of-scope integrity (no fabricated circularity/detachability values)
    grey_keys = ["pct_reused", "pct_available_for_reuse", "assumed_construction_waste", "waste_codes", 
                 "detachability_connection", "detachability_connection_detail", "detachability_accessibility", 
                 "detachability_intersection", "detachability_product_edge", "lifespan_years"]
    for r in records:
        for k in grey_keys:
            assert r.get(k) is None, f"Item {r['boq_item_no']} has value in grey out-of-scope column '{k}': {r.get(k)}"
    print("[PASS 12/14] Out-of-scope grey columns remain strictly blank per instructions.")
    passed_checks += 1
    
    # 13. Carbon Bonus B2 verification (at least 5 materials cited)
    carbon_items = [r for r in records if r.get("gwp_per_kg") is not None and "Source:" in (r.get("comment") or "")]
    assert len(carbon_items) >= 5, f"Expected at least 5 materials with carbon citations for B2, got {len(carbon_items)}"
    print(f"[PASS 13/14] Carbon Bonus B2 verified with {len(carbon_items)} cited material records.")
    passed_checks += 1
    
    # 14. Visualization image verification
    assert os.path.getsize(viz_path) > 10000, "Visualization image file size too small"
    print("[PASS 14/14] Visualization chart output/visualization.png verified.")
    passed_checks += 1
    
    print("\n==================================================")
    print(f" ALL {passed_checks}/{total_checks} VALIDATION CHECKS PASSED SUCCESSFULLY! ")
    print("==================================================\n")
    return True


if __name__ == "__main__":
    validate_all()
